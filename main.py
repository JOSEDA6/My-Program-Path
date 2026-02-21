# -*- coding: utf-8 -*-
import sys
import os
import re
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
from PySide6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout,
                               QHBoxLayout, QPushButton, QLabel, QListWidget,
                               QFileDialog, QProgressBar, QTextEdit, QMessageBox,
                               QListWidgetItem, QLineEdit)
from PySide6.QtCore import Qt, QThread, Signal, QSettings
from PySide6.QtGui import QFont, QIcon


def get_resource_path(relative_path):
    """获取资源绝对路径，适配 PyInstaller 单文件打包的临时目录"""
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.abspath("."), relative_path)


# ============================
# 配置与工具函数
# ============================
OUTPUT_FILENAME_PREFIX = "合并工单_"
UNIT_MAP = {
    "個": "个", "個/pcs": "个", "臺": "台", "臺/台": "台",
    "公斤": "kg", "千克": "kg", "g": "g", "公斤/公斤": "kg"
}


def safe_float(x):
    try:
        if x is None or str(x).strip() == '': return 0.0
        s = str(x).strip().replace(",", "")
        return float(s)
    except:
        return 0.0


# ============================
# 核心处理类：OrderProcessor
# ============================
class OrderProcessor:
    def __init__(self):
        self.standard_columns = ['序号', '品名', '规格/图号', '单位', '数量', '单价', '金额', '备注/本体单重']
        self.mapping_keywords = {
            '品名': ['品名', '物料名称', 'product name', 'material name'],
            '规格/图号': ['规格', '图号', '物料规格', 'spec', 'specification', '型号'],
            '单位': ['单位', 'unit', 'uom', '采购单位'],
            '数量': ['数量', 'qty', 'quantity', '采购数量', '报价数量'],
            '单价': ['单价', 'price', 'unit price', '报价单价'],
            '金额': ['金额', 'total', 'amount'],
            '备注/本体单重': ['备注', 'remarks', '本体单重', 'item no. remarks', '询价说明'],
            '询价人': ['询价人', 'inquirer'],
            '代购厂商': ['代购厂商', 'purchasing agent', '代购']
        }
        self.order_no_pattern = re.compile(r'XIDP-[A-Z]?\d{10,12}', re.I)

    def read_excel_smart(self, file_path):
        try:
            if file_path.endswith('.csv'):
                for enc in ['utf-8', 'gbk', 'gb18030', 'utf-8-sig']:
                    try:
                        return pd.read_csv(file_path, header=None, encoding=enc, dtype=str).fillna("")
                    except:
                        continue
            return pd.read_excel(file_path, header=None, dtype=str).fillna("")
        except:
            return pd.DataFrame()

    def parse_file_to_sections(self, file_path):
        df = self.read_excel_smart(file_path)
        if df.empty: return []

        global_id = "未知单号"
        for r in range(len(df)):
            row_str = " ".join([str(v) for v in df.iloc[r]])
            match = self.order_no_pattern.search(row_str)
            if match:
                global_id = match.group(0)
                break

        sections = []
        current_section = None
        header_map = None

        for r in range(len(df)):
            row_values = [str(df.iat[r, c]).strip() for c in range(len(df.columns))]
            row_str = " ".join(row_values)

            # 探测表头
            if any(k in row_str for k in ['品名', '物料名称', '规格']) and any(k in row_str for k in ['数量', '单价']):
                header_map = self.create_header_map(row_values)
                if current_section: current_section['header_map'] = header_map
                continue

            id_match = self.order_no_pattern.search(row_str)
            found_id = id_match.group(0) if id_match else None

            if found_id and (not current_section or found_id != current_section['order_no']):
                if current_section and not current_section['data_rows'].empty:
                    sections.append(current_section)

                date_match = re.search(r'(\d{4}[-/]\d{1,2}[-/]\d{1,2})', row_str)
                current_section = {
                    'order_no': found_id,
                    'date': date_match.group(1) if date_match else "",
                    'info': "",
                    'header_map': header_map,
                    'data_rows': pd.DataFrame(columns=self.standard_columns)
                }

            if header_map:
                item_data = self.map_row_to_std(row_values, header_map)
                p_name = item_data.get('品名', '')

                if p_name and p_name not in ['品名', '物料名称', 'Material Name', '物料名称(品名)']:
                    if not current_section:
                        current_section = {'order_no': global_id, 'date': "", 'info': "", 'header_map': header_map,
                                           'data_rows': pd.DataFrame(columns=self.standard_columns)}

                    # 提取询价信息
                    if not current_section['info']:
                        agent = str(item_data.get('代购厂商', '')).strip()
                        inquirer = str(item_data.get('询价人', '')).strip()
                        for empty in ['无', '無', 'none', 'null', '-', 'nan']:
                            if agent.lower() == empty: agent = ''
                            if inquirer.lower() == empty: inquirer = ''
                        parts = [p for p in [agent, inquirer] if p]
                        current_section['info'] = f"{'，'.join(parts)}" if parts else "工单详情"

                    qty = safe_float(item_data.get('数量'))
                    prc = safe_float(item_data.get('单价'))
                    item_data['数量'] = int(qty) if qty.is_integer() else qty
                    item_data['单价'] = round(prc, 2)
                    item_data['金额'] = round(qty * prc, 2)

                    filtered_item = {col: item_data.get(col, '') for col in self.standard_columns}
                    current_section['data_rows'] = pd.concat(
                        [current_section['data_rows'], pd.DataFrame([filtered_item])], ignore_index=True)

        if current_section and not current_section['data_rows'].empty:
            sections.append(current_section)
        return sections

    def create_header_map(self, row_values):
        h_map = {}
        for idx, val in enumerate(row_values):
            low_val = val.lower()
            for std_key, keywords in self.mapping_keywords.items():
                if std_key not in h_map and any(kw in low_val for kw in keywords):
                    h_map[std_key] = idx
        return h_map

    def map_row_to_std(self, row_values, h_map):
        res = {}
        for k, idx in h_map.items():
            if idx < len(row_values):
                val = row_values[idx]
                res[k] = self.normalize_unit(val) if k == '单位' else val
        return res

    def normalize_unit(self, u):
        for k, v in UNIT_MAP.items():
            if k in u: return v
        return u


# ============================
# 执行线程 WorkerThread
# ============================
class WorkerThread(QThread):
    progress_signal = Signal(int, str, int)
    log_signal = Signal(str)
    finished_signal = Signal(str, list)
    stopped_signal = Signal()

    def __init__(self, files, output_dir):
        super().__init__()
        self.files, self.output_dir = files, output_dir
        self.processor = OrderProcessor()
        self._abort = False

    def stop(self):
        self._abort = True

    def run(self):
        try:
            seen_rows = set()
            wb = Workbook()
            ws = wb.active
            ws.title = "汇总工单"
            thin, thick = Side(border_style="thin"), Side(border_style="medium")
            center = Alignment(horizontal="center", vertical="center", wrap_text=True)
            left_center = Alignment(horizontal="left", vertical="center", wrap_text=True)  # <--- 新增对齐样式
            bold_font = Font(bold=True)

            # 表头
            for i, name in enumerate(self.processor.standard_columns, 1):
                cell = ws.cell(1, i, name)
                cell.font, cell.alignment = bold_font, center
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

            curr_row = 2
            for idx, fpath in enumerate(self.files):
                if self._abort: break
                sections = self.processor.parse_file_to_sections(fpath)
                for section in sections:
                    start_row = curr_row
                    # 信息行写入与对齐
                    c1 = ws.cell(curr_row, 1, section['date'])
                    c1.alignment = center  # 日期居中

                    c2 = ws.cell(curr_row, 2, section['info'])
                    c2.alignment = left_center  # 询价人靠左、垂直居中

                    c3 = ws.cell(curr_row, 3, section['order_no'])
                    c3.alignment = center  # 单号居中

                    for c in range(1, 9): ws.cell(curr_row, c).border = Border(top=thin, left=thin, right=thin,
                                                                               bottom=thin)
                    curr_row += 1

                    written_count = 0
                    for _, row in section['data_rows'].iterrows():
                        sig = (section['order_no'], row['品名'], row['规格/图号'], str(row['数量']))
                        if sig in seen_rows: continue
                        seen_rows.add(sig)

                        for col_idx, col_name in enumerate(self.processor.standard_columns, 1):
                            val = row[col_name]
                            if col_name == '序号': val = written_count + 1
                            cell = ws.cell(curr_row, col_idx, val)
                            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                            cell.alignment = center if col_name in ['序号', '单位', '数量'] else Alignment(
                                vertical="center", horizontal="left")
                            if col_name in ['单价', '金额']: cell.number_format = '0.00'
                        curr_row += 1
                        written_count += 1

                    if written_count == 0:
                        curr_row -= 1
                    else:
                        self.apply_outer_border(ws, start_row, curr_row - 1, thick, thin)
                        curr_row += 1

                self.progress_signal.emit(int((idx + 1) / len(self.files) * 100), os.path.basename(fpath), idx)

            if curr_row > 2:
                out = os.path.join(self.output_dir,
                                   f"{OUTPUT_FILENAME_PREFIX}{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
                wb.save(out)
                self.finished_signal.emit(out, self.files)
            else:
                self.stopped_signal.emit()
        except Exception as e:
            self.log_signal.emit(f"❌ 错误：{str(e)}")
            self.stopped_signal.emit()

    def apply_outer_border(self, ws, start_r, end_r, thick, thin):
        for r in range(start_r, end_r + 1):
            for c in range(1, 9):
                l = thick if c == 1 else thin
                r_s = thick if c == 8 else thin
                t = thick if r == start_r else thin
                b = thick if r == end_r else thin
                ws.cell(r, c).border = Border(left=l, right=r_s, top=t, bottom=b)


# ============================
# GUI 保持不变 (MainWindow & Style)
# ============================
STYLE = """
QMainWindow { background: #F5F6F8; }
QLabel { font-size: 14px; color: #1C1C1E; }
QListWidget { border: 1px solid #E6E6EA; border-radius: 8px; background: white; }
QTextEdit { background: #0F0F10; color: #9EE39A; border-radius: 8px; font-family: 'Consolas'; font-size: 11px; }
QPushButton#Run { background: #FF9500; color: white; border-radius: 10px; height: 40px; font-weight: 600; font-size: 15px; }
QPushButton#Run:disabled { background: #D9D9DC; color: #888888; }
QPushButton#Action { background: white; border: 1px solid #E6E6EA; border-radius: 8px; height: 32px; padding: 0 15px; }
QProgressBar { background: #ECECF0; border-radius: 6px; height: 12px; text-align: center; font-size: 10px; }
QProgressBar::chunk { background: #FF9500; border-radius: 6px; }
"""


class FileItemWidget(QWidget):
    def __init__(self, file_path, remove_callback):
        super().__init__()
        self.file_path = file_path
        h = QHBoxLayout(self)
        h.setContentsMargins(10, 5, 10, 5)
        lbl = QLabel(os.path.basename(file_path))
        btn = QPushButton("移除")
        btn.setFixedSize(60, 24)
        btn.setObjectName("Action")
        btn.clicked.connect(lambda: remove_callback(file_path))
        h.addWidget(lbl)
        h.addStretch()
        h.addWidget(btn)


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("工单汇总工具")
        self.resize(1000, 750)
        self.setStyleSheet(STYLE)
        self.settings = QSettings("MySoft", "OrderMerge")
        self.output_dir = self.settings.value("output_dir", os.path.join(os.path.expanduser("~"), "Desktop"))
        self.files = []
        self.init_ui()

    def init_ui(self):
        w = QWidget()
        self.setCentralWidget(w)
        lay = QVBoxLayout(w)
        lay.setContentsMargins(25, 20, 25, 20)
        header = QHBoxLayout()
        header.addWidget(QLabel("待处理文件："))
        header.addStretch()
        btn_add = QPushButton("添加 Excel/CSV")
        btn_add.setObjectName("Action")
        btn_add.clicked.connect(self.add_files)
        btn_clear = QPushButton("清空")
        btn_clear.setObjectName("Action")
        btn_clear.clicked.connect(self.clear_list)
        header.addWidget(btn_add)
        header.addWidget(btn_clear)
        lay.addLayout(header)
        lists = QHBoxLayout()
        self.file_list = QListWidget()
        lists.addWidget(self.file_list, 3)
        log_v = QVBoxLayout()
        log_v.addWidget(QLabel("执行日志："))
        self.log = QTextEdit()
        self.log.setReadOnly(True)
        log_v.addWidget(self.log)
        lists.addLayout(log_v, 2)
        lay.addLayout(lists)
        out_lay = QHBoxLayout()
        self.output_edit = QLineEdit(self.output_dir)
        self.output_edit.setReadOnly(True)
        btn_dir = QPushButton("修改保存位置")
        btn_dir.setObjectName("Action")
        btn_dir.clicked.connect(self.choose_output_dir)
        out_lay.addWidget(QLabel("保存到："))
        out_lay.addWidget(self.output_edit)
        out_lay.addWidget(btn_dir)
        lay.addLayout(out_lay)
        lay.addSpacing(10)
        self.pbar = QProgressBar()
        lay.addWidget(self.pbar)
        ops = QHBoxLayout()
        self.btn_run = QPushButton("开始合并任务")
        self.btn_run.setObjectName("Run")
        self.btn_run.clicked.connect(self.start_merge)
        self.btn_stop = QPushButton("终止")
        self.btn_stop.setObjectName("Action")
        self.btn_stop.setEnabled(False)
        self.btn_stop.clicked.connect(self.stop_merge)
        ops.addWidget(self.btn_run)
        ops.addWidget(self.btn_stop)
        ops.addStretch()
        lay.addLayout(ops)

    def add_files(self):
        paths, _ = QFileDialog.getOpenFileNames(self, "选择文件", "", "Excel/CSV (*.xls *.xlsx *.csv)")
        for p in paths:
            if p not in self.files:
                self.files.append(p)
                item = QListWidgetItem(self.file_list)
                widget = FileItemWidget(p, self.remove_file)
                item.setSizeHint(widget.sizeHint())
                self.file_list.setItemWidget(item, widget)

    def remove_file(self, path):
        if path in self.files: self.files.remove(path)
        for i in range(self.file_list.count()):
            w = self.file_list.itemWidget(self.file_list.item(i))
            if w and w.file_path == path:
                self.file_list.takeItem(i)
                break

    def clear_list(self):
        self.files = []
        self.file_list.clear()

    def choose_output_dir(self):
        d = QFileDialog.getExistingDirectory(self, "选择目录", self.output_dir)
        if d:
            self.output_dir = d
            self.output_edit.setText(d)
            self.settings.setValue("output_dir", d)

    def start_merge(self):
        if not self.files: return
        self.log.clear()
        self.pbar.setValue(0)
        self.btn_run.setEnabled(False)
        self.btn_stop.setEnabled(True)
        self.worker = WorkerThread(self.files, self.output_dir)
        self.worker.progress_signal.connect(lambda v, n, i: (self.pbar.setValue(v), self.file_list.setCurrentRow(i)))
        self.worker.log_signal.connect(lambda t: self.log.append(f"[{datetime.now().strftime('%H:%M:%S')}] {t}"))
        self.worker.finished_signal.connect(self.on_finished)
        self.worker.stopped_signal.connect(self.on_stopped)
        self.worker.start()

    def stop_merge(self):
        if hasattr(self, 'worker'): self.worker.stop()

    def on_finished(self, path, logs):
        QMessageBox.information(self, "完成", f"任务已完成！\n保存至：{path}")
        self.reset_ui()
        os.startfile(os.path.dirname(path))

    def on_stopped(self):
        self.reset_ui()

    def reset_ui(self):
        self.btn_run.setEnabled(True)
        self.btn_stop.setEnabled(False)


if __name__ == "__main__":
    import multiprocessing

    multiprocessing.freeze_support()
    app = QApplication(sys.argv)
    app.setWindowIcon(QIcon(get_resource_path("app.ico")))
    font = app.font()
    font.setFamily("Microsoft YaHei")
    font.setPointSize(10)
    app.setFont(font)
    win = MainWindow()
    win.show()
    sys.exit(app.exec())